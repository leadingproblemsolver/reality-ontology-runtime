from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable


class ConstructionInputError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class SourceRef:
    path: str
    row: int
    digest: str


@dataclass(frozen=True, slots=True)
class Activity:
    activity_id: str
    activity_name: str
    wbs: str
    planned_start: date
    planned_finish: date
    status: str
    source: SourceRef


@dataclass(frozen=True, slots=True)
class ActivityRelation:
    predecessor_id: str
    successor_id: str
    source: SourceRef


@dataclass(frozen=True, slots=True)
class Requirement:
    scope_type: str
    scope_id: str
    requirement: str
    lead_days: int
    source: SourceRef


@dataclass(frozen=True, slots=True)
class ProjectionRow:
    activity_id: str
    activity_name: str
    wbs: str
    planned_start: str
    planned_finish: str
    activity_status: str
    readiness_status: str
    blocking_predecessors: str
    requirement: str
    action_by: str
    source_activity: str
    source_requirement: str


ALIASES = {
    "activity_id": {"activity_id", "activity id", "activity id*", "id"},
    "activity_name": {"activity_name", "activity name", "name"},
    "wbs": {"wbs", "wbs code", "wbs_name", "wbs name"},
    "planned_start": {"planned_start", "planned start", "start", "planned start date"},
    "planned_finish": {"planned_finish", "planned finish", "finish", "planned finish date"},
    "status": {"status", "activity status"},
    "predecessor_id": {"predecessor_id", "predecessor id", "predecessor"},
    "successor_id": {"successor_id", "successor id", "successor"},
    "scope_type": {"scope_type", "scope type", "applies to type"},
    "scope_id": {"scope_id", "scope id", "applies to", "applies to id"},
    "requirement": {"requirement", "documentation", "document", "inspection requirement"},
    "lead_days": {"lead_days", "lead days", "prepare days before"},
}


def _norm(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _canonical_headers(headers: Iterable[str]) -> dict[str, str]:
    resolved: dict[str, str] = {}
    normalized = {_norm(h): h for h in headers if h is not None}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            if _norm(alias) in normalized:
                resolved[canonical] = normalized[_norm(alias)]
                break
    return resolved


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_rows(path: str | Path, sheet: str | None = None) -> tuple[list[dict[str, object]], str]:
    p = Path(path)
    if not p.exists():
        raise ConstructionInputError(f"source does not exist: {p}")
    digest = _digest(p)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        with p.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.DictReader(fh))
        return rows, digest
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ConstructionInputError(
                "XLSX input requires openpyxl; install with `pip install .[xlsx]`"
            ) from exc
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = [str(v or "").strip() for v in next(iterator)]
        except StopIteration:
            return [], digest
        rows = [dict(zip(headers, values)) for values in iterator]
        return rows, digest
    raise ConstructionInputError(f"unsupported source type: {suffix}; use CSV or XLSX")


def _parse_date(value: object, *, field: str, row: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ConstructionInputError(f"invalid {field} at row {row}: {text!r}")


def load_activities(path: str | Path, *, sheet: str | None = None) -> list[Activity]:
    rows, digest = _read_rows(path, sheet)
    if not rows:
        return []
    headers = _canonical_headers(rows[0].keys())
    required = {"activity_id", "activity_name", "wbs", "planned_start", "planned_finish", "status"}
    missing = sorted(required - headers.keys())
    if missing:
        raise ConstructionInputError(f"activity source missing columns: {', '.join(missing)}")
    activities: list[Activity] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows, 2):
        activity_id = str(row.get(headers["activity_id"], "") or "").strip()
        if not activity_id:
            continue
        if activity_id in seen:
            raise ConstructionInputError(f"duplicate activity_id {activity_id!r} at row {row_number}")
        seen.add(activity_id)
        start = _parse_date(row.get(headers["planned_start"]), field="planned_start", row=row_number)
        finish = _parse_date(row.get(headers["planned_finish"]), field="planned_finish", row=row_number)
        if finish < start:
            raise ConstructionInputError(f"planned_finish before planned_start for {activity_id}")
        activities.append(
            Activity(
                activity_id=activity_id,
                activity_name=str(row.get(headers["activity_name"], "") or "").strip(),
                wbs=str(row.get(headers["wbs"], "") or "").strip(),
                planned_start=start,
                planned_finish=finish,
                status=str(row.get(headers["status"], "") or "").strip(),
                source=SourceRef(str(Path(path)), row_number, digest),
            )
        )
    return activities


def load_relations(path: str | Path, *, sheet: str | None = None) -> list[ActivityRelation]:
    rows, digest = _read_rows(path, sheet)
    if not rows:
        return []
    headers = _canonical_headers(rows[0].keys())
    required = {"predecessor_id", "successor_id"}
    missing = sorted(required - headers.keys())
    if missing:
        raise ConstructionInputError(f"relationship source missing columns: {', '.join(missing)}")
    out = []
    for row_number, row in enumerate(rows, 2):
        predecessor = str(row.get(headers["predecessor_id"], "") or "").strip()
        successor = str(row.get(headers["successor_id"], "") or "").strip()
        if predecessor and successor:
            out.append(ActivityRelation(predecessor, successor, SourceRef(str(Path(path)), row_number, digest)))
    return out


def load_requirements(path: str | Path, *, sheet: str | None = None) -> list[Requirement]:
    rows, digest = _read_rows(path, sheet)
    if not rows:
        return []
    headers = _canonical_headers(rows[0].keys())
    required = {"scope_type", "scope_id", "requirement"}
    missing = sorted(required - headers.keys())
    if missing:
        raise ConstructionInputError(f"requirements source missing columns: {', '.join(missing)}")
    out = []
    for row_number, row in enumerate(rows, 2):
        scope_type = _norm(row.get(headers["scope_type"])).replace(" ", "_")
        scope_id = str(row.get(headers["scope_id"], "") or "").strip()
        requirement = str(row.get(headers["requirement"], "") or "").strip()
        if not scope_id or not requirement:
            continue
        if scope_type not in {"activity", "wbs"}:
            raise ConstructionInputError(f"unsupported scope_type {scope_type!r} at row {row_number}")
        lead_days = 0
        if "lead_days" in headers:
            raw = row.get(headers["lead_days"])
            if str(raw or "").strip():
                try:
                    lead_days = int(float(str(raw)))
                except ValueError as exc:
                    raise ConstructionInputError(f"invalid lead_days at row {row_number}: {raw!r}") from exc
        out.append(Requirement(scope_type, scope_id, requirement, lead_days, SourceRef(str(Path(path)), row_number, digest)))
    return out


def _is_complete(status: str) -> bool:
    return _norm(status) in {"complete", "completed", "finished", "actual finish"}


def _source_label(ref: SourceRef) -> str:
    return f"{ref.path}#row={ref.row}@sha256:{ref.digest[:12]}"


def project_lookahead(
    activities: list[Activity],
    relations: list[ActivityRelation],
    requirements: list[Requirement],
    *,
    as_of: date,
    days: int = 90,
) -> list[ProjectionRow]:
    if days <= 0:
        raise ConstructionInputError("days must be > 0")
    by_id = {a.activity_id: a for a in activities}
    unknown_relations = [r for r in relations if r.predecessor_id not in by_id or r.successor_id not in by_id]
    if unknown_relations:
        first = unknown_relations[0]
        raise ConstructionInputError(
            f"relationship references unknown activity: {first.predecessor_id}->{first.successor_id}"
        )
    predecessors: dict[str, list[str]] = {}
    for relation in relations:
        predecessors.setdefault(relation.successor_id, []).append(relation.predecessor_id)

    req_by_activity: dict[str, list[Requirement]] = {}
    req_by_wbs: dict[str, list[Requirement]] = {}
    for req in requirements:
        target = req_by_activity if req.scope_type == "activity" else req_by_wbs
        target.setdefault(req.scope_id, []).append(req)

    horizon = as_of + timedelta(days=days)
    rows: list[ProjectionRow] = []
    for activity in sorted(activities, key=lambda a: (a.planned_start, a.activity_id)):
        if _is_complete(activity.status):
            continue
        if activity.planned_start > horizon or activity.planned_finish < as_of:
            continue
        blockers = [
            pred_id for pred_id in predecessors.get(activity.activity_id, []) if not _is_complete(by_id[pred_id].status)
        ]
        matched = req_by_activity.get(activity.activity_id) or req_by_wbs.get(activity.wbs) or []
        if not matched:
            rows.append(
                ProjectionRow(
                    activity.activity_id,
                    activity.activity_name,
                    activity.wbs,
                    activity.planned_start.isoformat(),
                    activity.planned_finish.isoformat(),
                    activity.status,
                    "BLOCKED" if blockers else "REQUIREMENT_UNMAPPED",
                    ";".join(blockers),
                    "",
                    "",
                    _source_label(activity.source),
                    "",
                )
            )
            continue
        for req in matched:
            action_by = activity.planned_start - timedelta(days=req.lead_days)
            if blockers:
                readiness = "BLOCKED"
            elif action_by <= as_of:
                readiness = "PREPARE_NOW"
            else:
                readiness = "UPCOMING"
            rows.append(
                ProjectionRow(
                    activity.activity_id,
                    activity.activity_name,
                    activity.wbs,
                    activity.planned_start.isoformat(),
                    activity.planned_finish.isoformat(),
                    activity.status,
                    readiness,
                    ";".join(blockers),
                    req.requirement,
                    action_by.isoformat(),
                    _source_label(activity.source),
                    _source_label(req.source),
                )
            )
    return rows


def diff_activities(previous: list[Activity], current: list[Activity], *, as_of: date, days: int = 90) -> list[dict[str, object]]:
    old = {a.activity_id: a for a in previous}
    new = {a.activity_id: a for a in current}
    horizon = as_of + timedelta(days=days)

    def in_window(a: Activity) -> bool:
        return not _is_complete(a.status) and a.planned_start <= horizon and a.planned_finish >= as_of

    changes: list[dict[str, object]] = []
    for activity_id in sorted(old.keys() | new.keys()):
        before, after = old.get(activity_id), new.get(activity_id)
        if before is None:
            changes.append({"activity_id": activity_id, "change": "NEW_ACTIVITY", "before": "", "after": after.activity_name})
            continue
        if after is None:
            changes.append({"activity_id": activity_id, "change": "REMOVED_ACTIVITY", "before": before.activity_name, "after": ""})
            continue
        if before.planned_start != after.planned_start:
            changes.append({
                "activity_id": activity_id,
                "change": "START_MOVED",
                "before": before.planned_start.isoformat(),
                "after": after.planned_start.isoformat(),
                "delta_days": (after.planned_start - before.planned_start).days,
            })
        if before.planned_finish != after.planned_finish:
            changes.append({
                "activity_id": activity_id,
                "change": "FINISH_MOVED",
                "before": before.planned_finish.isoformat(),
                "after": after.planned_finish.isoformat(),
                "delta_days": (after.planned_finish - before.planned_finish).days,
            })
        if _norm(before.status) != _norm(after.status):
            changes.append({"activity_id": activity_id, "change": "STATUS_CHANGED", "before": before.status, "after": after.status})
        if in_window(before) != in_window(after):
            changes.append({
                "activity_id": activity_id,
                "change": "ENTERED_LOOKAHEAD" if in_window(after) else "LEFT_LOOKAHEAD",
                "before": str(in_window(before)),
                "after": str(in_window(after)),
            })
    return changes


def build_change_impact(
    changes: list[dict[str, object]],
    projection: list[ProjectionRow],
    previous: list[Activity],
    current: list[Activity],
) -> list[dict[str, object]]:
    previous_by_id = {activity.activity_id: activity for activity in previous}
    current_by_id = {activity.activity_id: activity for activity in current}
    readiness_by_activity: dict[str, list[ProjectionRow]] = {}
    for row in projection:
        readiness_by_activity.setdefault(row.activity_id, []).append(row)
    for rows in readiness_by_activity.values():
        rows.sort(key=lambda row: (row.requirement, row.readiness_status, row.action_by, row.source_requirement))

    impacts: list[dict[str, object]] = []
    for change in changes:
        activity_id = str(change["activity_id"])
        before_activity = previous_by_id.get(activity_id)
        after_activity = current_by_id.get(activity_id)
        source_before = _source_label(before_activity.source) if before_activity else ""
        source_after = _source_label(after_activity.source) if after_activity else ""
        activity = after_activity or before_activity
        base = {
            "activity_id": activity_id,
            "change": change["change"],
            "before": change.get("before", ""),
            "after": change.get("after", ""),
            "delta_days": change.get("delta_days", ""),
            "source_before": source_before,
            "source_after": source_after,
        }
        downstream = readiness_by_activity.get(activity_id, [])
        if not downstream:
            impacts.append({
                **base,
                "activity_name": activity.activity_name if activity else "",
                "wbs": activity.wbs if activity else "",
                "readiness_status": "",
                "blocking_predecessors": "",
                "requirement": "",
                "action_by": "",
                "effect_evidence": "UNRESOLVED",
                "source_activity": source_after,
                "source_requirement": "",
            })
            continue
        for row in downstream:
            impacts.append({
                **base,
                "activity_name": row.activity_name,
                "wbs": row.wbs,
                "readiness_status": row.readiness_status,
                "blocking_predecessors": row.blocking_predecessors,
                "requirement": row.requirement,
                "action_by": row.action_by,
                "effect_evidence": "KNOWN" if row.source_requirement else "UNRESOLVED",
                "source_activity": row.source_activity,
                "source_requirement": row.source_requirement,
            })
    return impacts


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def run_lookahead(
    *,
    activities_path: str,
    relationships_path: str,
    requirements_path: str,
    as_of: date,
    days: int = 90,
    output_dir: str = "artifacts/construction",
    previous_activities_path: str | None = None,
    with_impact: bool = False,
) -> dict[str, object]:
    if with_impact and not previous_activities_path:
        raise ConstructionInputError("--with-impact requires --previous-activities")

    activities = load_activities(activities_path)
    relations = load_relations(relationships_path)
    requirements = load_requirements(requirements_path)
    projection = project_lookahead(activities, relations, requirements, as_of=as_of, days=days)

    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    readiness_path = target / f"{days}_day_readiness.csv"
    _write_csv(readiness_path, [asdict(row) for row in projection])

    changes: list[dict[str, object]] = []
    changes_path = target / "schedule_changes.csv"
    impact: list[dict[str, object]] = []
    impact_path = target / "change_impact.csv"
    if previous_activities_path:
        previous = load_activities(previous_activities_path)
        changes = diff_activities(previous, activities, as_of=as_of, days=days)
        _write_csv(changes_path, changes)
        if with_impact:
            impact = build_change_impact(changes, projection, previous, activities)
            _write_csv(impact_path, impact)

    source_hashes = {
        "activities": _digest(Path(activities_path)),
        "relationships": _digest(Path(relationships_path)),
        "requirements": _digest(Path(requirements_path)),
    }
    if previous_activities_path:
        source_hashes["previous_activities"] = _digest(Path(previous_activities_path))

    receipt = {
        "projection_type": "construction_lookahead_readiness",
        "as_of": as_of.isoformat(),
        "window_days": days,
        "source_hashes": source_hashes,
        "activities_loaded": len(activities),
        "relations_loaded": len(relations),
        "requirements_loaded": len(requirements),
        "projection_rows": len(projection),
        "blocked_rows": sum(1 for r in projection if r.readiness_status == "BLOCKED"),
        "unmapped_rows": sum(1 for r in projection if r.readiness_status == "REQUIREMENT_UNMAPPED"),
        "changes": len(changes),
        "outputs": {
            "readiness_csv": str(readiness_path),
            "changes_csv": str(changes_path) if previous_activities_path else None,
        },
    }
    if with_impact:
        receipt["impact_rows"] = len(impact)
        receipt["outputs"]["impact_csv"] = str(impact_path)
    receipt_text = json.dumps(receipt, sort_keys=True, indent=2)
    receipt["projection_digest"] = hashlib.sha256(receipt_text.encode()).hexdigest()
    receipt_path = target / "projection_receipt.json"
    receipt_path.write_text(json.dumps(receipt, indent=2), encoding="utf-8")
    receipt["outputs"]["receipt"] = str(receipt_path)
    return receipt
